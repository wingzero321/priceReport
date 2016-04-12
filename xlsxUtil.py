# -*- coding: utf-8 -*-

import sys
from openpyxl.styles import *
from style import Style
import xlsxFormatSetting as settings
import cx_Oracle

reload(sys)
sys.setdefaultencoding('utf8')
reload(sys)

class XlsxUtil:
    def __init__(self, sheet):
        self.sheet = sheet

    #生成"报表使用说明"
    def createReportInstruction(self, dateList, colNum, rowNum):
        #表格各部分高度参数
        colHeight = {'title': 21, 'thead': 19.5, 'content': 35.25}

        thead = ['分析报表', '分析时间段', '筛选条件', '分析目的']
        #标题"报表使用说明"应该与表头等宽
        self.mergeRowCell('报表使用说明', colNum, rowNum, len(thead), colHeight['title'], Style('标题'))

        rowNum += 1
        #设置表头
        self.setRow(colNum, rowNum, thead, Style('行标题'), colHeight['thead'])

        if len(dateList) == 1:
            timeStr = dateList[0].strftime('%m-%d')
        else:
            timeStr = dateList[0].strftime('%m-%d') + '~' + dateList[len(dateList)-1].strftime('%m-%d')

        rowNum += 1
        row = ['网站经销商数据变化分析', timeStr, '无', '分析网站报价经销商数变化情况，发现数量变化突跳']
        self.setRow(colNum, rowNum, row, Style('内容'), colHeight['content'])

        rowNum += 1
        row = ['车型网站报价均值分析', timeStr, '无', '分析网站之间价格差异']
        self.setRow(colNum, rowNum, row, Style('内容'), colHeight['content'])

        rowNum += 1
        row = ['大区报价分析', timeStr, '4S店,直营店,卫星店', '分析大区之间的异常报价与显著差异']
        self.setRow(colNum, rowNum, row, Style('内容'), colHeight['content'])

        rowNum += 1
        row = ['省份报价分析', timeStr, '4S店,直营店,卫星店', '分析大区下省份之间的异常报价与显著差异']
        self.setRow(colNum, rowNum, row, Style('内容'), colHeight['content'])

        rowNum += 1
        row = ['报价信息', timeStr, '无', '一天内的报价详细信息']
        self.setRow(colNum, rowNum, row, Style('内容'), colHeight['content'])

        return rowNum+1

    def createReportParameterSetting(self, conn, series, dateList, colNum, rowNum):
        colHeight = {'title': 24.75, 'thead': 25.5, 'content': 25.5}

        # thead = ['范围', '车型', '平均值 网络报价', '平均值 MSRP', '平均值 差价', '标杆值', '报警原则']
        thead = ['范围', '车型', '平均值 网络报价', '平均值 MSRP', '平均值 差价', '标杆值']
        self.mergeRowCell('报表参数设置【（参数-报价）应用于非大区分析报表 - 突出显示】', colNum, rowNum,
                          len(thead), colHeight['title'], Style('标题'))

        rowNum += 1
        self.setRow(colNum, rowNum, thead, Style(), colHeight['thead'])

        #最后两个单元格加粗,字体改为宋体
        self.setCellBold(self.sheet[self.parseCellID(colNum+len(thead)-1, rowNum)])
        self.setCellSongTi(self.sheet[self.parseCellID(colNum+len(thead), rowNum)])
        self.setCellBold(self.sheet[self.parseCellID(colNum+len(thead)-1, rowNum)])
        self.setCellSongTi(self.sheet[self.parseCellID(colNum+len(thead), rowNum)])
        #标杆值一栏改变背景色
        self.setCellBgColor(self.sheet[self.parseCellID(colNum+len(thead)-1, rowNum)], 'FFC000')

        dateList[0].strftime('%m-%d')
        values = []
        sqlTimeStr = ""
        for i in range(len(dateList)):
            if i == 0:
                sqlTimeStr += ' and ( '
            else:
                sqlTimeStr += ' or '
            values.append(dateList[i].strftime('%Y-%m-%d'))
            sqlTimeStr += " crawl_date = :1 "
        sqlTimeStr += ' ) '
        values.append(series)


        cursor = conn.cursor()
        #result_all 查询4S店,直营店,卫星店的信息
        sql = 'select CARTYPE_NAME, avg(PRICE) PRICE, avg(MSRP) MSRP, avg(msrp)-avg(price) DIFFERENTIAL from price_report ' \
              'where IS_4S != \'#N/A\' ' + sqlTimeStr + ' and carseries_name = :1 group by CARTYPE_NAME'
        cursor.execute(sql, values)
        result_all = cursor.fetchall()
        result_allMap = {}
        for cell in result_all:
            if result_allMap.has_key(cell[0]):
                result_allMap[cell[0]]['price'] = cell[1]
                result_allMap[cell[0]]['msrp'] = cell[2]
                result_allMap[cell[0]]['differential'] = cell[3]
            else:
                result_allMap[cell[0]] = {}
                result_allMap[cell[0]]['price'] = cell[1]
                result_allMap[cell[0]]['msrp'] = cell[2]
                result_allMap[cell[0]]['differential'] = cell[3]

        #result_NA 查询未匹配经销商名称的信息
        sql = 'select CARTYPE_NAME, avg(PRICE) PRICE, avg(MSRP) MSRP, avg(msrp)-avg(price) DIFFERENTIAL from price_report ' \
              'where IS_4S = \'#N/A\' ' + sqlTimeStr + ' and carseries_name = :1  group by CARTYPE_NAME'
        cursor.execute(sql, values)
        result_NA = cursor.fetchall()
        result_NAMap = {}
        for cell in result_NA:
            if result_NAMap.has_key(cell[0]):
                result_NAMap[cell[0]]['price'] = cell[1]
                result_NAMap[cell[0]]['msrp'] = cell[2]
                result_NAMap[cell[0]]['differential'] = cell[3]
            else:
                result_NAMap[cell[0]] = {}
                result_NAMap[cell[0]]['price'] = cell[1]
                result_NAMap[cell[0]]['msrp'] = cell[2]
                result_NAMap[cell[0]]['differential'] = cell[3]
        cursor.close()

        rowNum += 1
        if len(result_all) + len(result_NA) <= 0:
            print '未查询到 报表参数设置 的数据'.encode('gbk')
            return -1
        else:
            if len(result_all) > 0:
                length_all = 0
                for cartype in self.sortCartype(conn, series):
                    if result_allMap.has_key(cartype):
                        rowValues = []
                        rowValues.append(cartype)
                        rowValues.append(result_allMap[cartype]['price'])
                        rowValues.append(result_allMap[cartype]['msrp'])
                        rowValues.append(result_allMap[cartype]['differential'])
                        rowValues.append(settings.markValue[series][rowValues[0]])
                        self.setRow(colNum+1, rowNum, rowValues, Style(), colHeight['content'])
                        #标杆值改背景色
                        self.setCellBgColor(self.sheet[self.parseCellID(colNum+len(rowValues), rowNum)], 'FFC000')
                        rowNum += 1
                        length_all += 1
                self.mergeColCell('4S店,直营店,卫星店', colNum, rowNum-length_all, length_all, Style())

            if len(result_NA) > 0:
                length_NA = 0
                for cartype in self.sortCartype(conn, series):
                    if result_NAMap.has_key(cartype):
                        rowValues = []
                        rowValues.append(cartype)
                        rowValues.append(result_NAMap[cartype]['price'])
                        rowValues.append(result_NAMap[cartype]['msrp'])
                        rowValues.append(result_NAMap[cartype]['differential'])
                        rowValues.append(settings.markValue[series][rowValues[0]])
                        self.setRow(colNum+1, rowNum, rowValues, Style(), colHeight['content'])
                        #标杆值改背景色
                        self.setCellBgColor(self.sheet[self.parseCellID(colNum+len(rowValues), rowNum)], 'FFC000')
                        rowNum += 1
                        length_NA += 1
                self.mergeColCell('4S店,直营店,卫星店', colNum, rowNum-length_NA, length_NA, Style())

            # self.mergeColCell('MSRP减报价均值,低于标杆值报红色预警', colNum+len(thead)-1, rowNum-(length_all+length_NA),
            #                   length_all+length_NA, Style())

        return rowNum

    def createReportsAnalysisOfWebsiteDealer(self, conn, series, dateList, colNum, rowNum, type):
        colHeight = {'title': 21, 'thead': 22.5, 'content': 22.5}

        if type == 'all':
            titleType = '4S店,直营店,卫星店'
        else:
            titleType = '未匹配经销商名称'
        self.mergeRowCell(titleType+' - 网站经销商数量分析', colNum, rowNum, len(dateList)+2, colHeight['title'], Style('标题'))

        rowNum += 1
        self.mergeRowCell('报价日期 日期', colNum+1, rowNum, len(dateList)+1, colHeight['title'], Style('行标题2'))

        rowNum += 1
        thead = ['网站来源']
        for cell in dateList:
            thead.append(cell.strftime('%m月%d日'))
        thead.append('总计')
        for i in range(9-len(thead)):
            thead.append('')
        self.setRow(colNum, rowNum, thead, Style(), colHeight['thead'])

        cursor = conn.cursor()
        values = [series]
        wheresql = ' where carseries_name = :1 '
        if type == 'all':
            wheresql += ' and IS_4S != \'#N/A\' '
        else:
            wheresql += ' and IS_4S = \'#N/A\' '
        wheresql += ' and ( '
        for i in range(len(dateList)):
            if i != 0:
                wheresql += ' or '
            values.append(dateList[i].strftime('%Y-%m-%d'))
            wheresql += ' crawl_date = :1 '
        wheresql += ' ) '
        sql = 'select SOURCE, crawl_date, COUNT(*) COUNT ' \
              'from ( select distinct SOURCE, crawl_date, NVL(dealerName,agency_name) from price_report ' + wheresql + \
              ' ) group by SOURCE, crawl_date'
        cursor.execute(sql, values)
        result = cursor.fetchall()

        sql = 'select SOURCE, COUNT(*) COUNT ' \
              'from ( select distinct SOURCE, NVL(dealerName,agency_name) from price_report ' + wheresql + \
              ' ) group by SOURCE'
        cursor.execute(sql, values)
        result_all = cursor.fetchall()
        if len(result) <= 0:
            print '网站经销商数量分析 未找到数据'.encode('gbk')
            return -1
        resultMap = {}
        for cell in result:
            resultMap[cell[0] + ' ' + cell[1]] = cell[2]

        rowNum += 1
        length = 0
        for rowValue in settings.websiteDealer:
            row = [rowValue]
            hasData = False
            for colValue in dateList:
                key = rowValue + ' ' + colValue.strftime('%Y-%m-%d')
                if resultMap.has_key(key):
                    row.append(resultMap[key])
                    hasData = True
                else:
                    row.append('')
            if hasData:
                for cell in result_all:
                    if cell[0] == rowValue:
                        row.append(cell[1])
                self.setRow(colNum, rowNum, row, Style(), colHeight['content'])
                rowNum += 1
                length += 1
        row = ['总计']

        sql = 'select crawl_date, COUNT(*) COUNT ' \
              'from ( select distinct crawl_date, NVL(dealerName,agency_name) from price_report ' + wheresql + \
              ' ) group by crawl_date'
        cursor.execute(sql, values)
        result_day = cursor.fetchall()

        sql = 'select COUNT(*) COUNT ' \
              'from ( select distinct NVL(dealerName,agency_name) from price_report ' + wheresql + \
              ' )'
        cursor.execute(sql, values)
        result_week = cursor.fetchall()
        cursor.close()

        for i in range(len(dateList)):
            for cell in result_day:
                if dateList[i].strftime('%Y-%m-%d') == cell[0]:
                    row.append(cell[1])
                    break
        row.append(result_week[0][0])
        self.setRow(colNum, rowNum, row, Style(), colHeight['content'])

        return rowNum

    def createAnalysisOfNetworkOffer(self, conn, series, dateList, colNum, rowNum, type):
        colHeight = {'title': 21, 'thead': 13.5, 'content': 13.5}

        if type == 'all':
            titleType = '4S店,直营店,卫星店'
            title_end = ''
        elif type == 'null':
            titleType = '未匹配经销商名称'
            title_end = ''
        elif type == 'all_autohome':
            titleType = '4S店,直营店,卫星店'
            title_end = '(汽车之家与易车)'
        else:
            titleType = '未匹配经销商名称'
            title_end = '(汽车之家与易车)'
        self.mergeRowCell(titleType+' - 网站报价均值分析'+title_end, colNum, rowNum,
                          len(dateList)+3, colHeight['title'], Style('标题'))

        rowNum += 1
        self.mergeRowCell('报价日期 日期', colNum+2, rowNum, len(dateList)+1, colHeight['title'], Style('行标题2'))

        rowNum += 1
        thead = ['车型', '网站来源']
        for i in range(len(dateList)):
            thead.append(dateList[i].strftime('%m月%d日'))
        thead.append('总计')
        self.setRow(colNum, rowNum, thead, Style(), colHeight['thead'])

        cursor = conn.cursor()
        values = [series]
        wheresql = ' where carseries_name = :1 '
        if type == 'all' or type == 'all_autohome':
            wheresql += ' and IS_4S != \'#N/A\' '
        else:
            wheresql += ' and IS_4S = \'#N/A\' '
        wheresql += ' and ( '
        for i in range(len(dateList)):
            if i != 0:
                wheresql += ' or '
            values.append(dateList[i].strftime('%Y-%m-%d'))
            wheresql += ' crawl_date = :1 '
        wheresql += ' ) '
        if type == 'all_autohome' or type == 'null_autohome':
            wheresql += ' and ( SOURCE like \'汽车之家%\' or  SOURCE like \'易车%\' ) '

        sql = 'select cartype_name, source, crawl_date, avg(msrp)-avg(price) differential ' \
              'from ( select cartype_name, source, price, msrp, crawl_date ' \
                        'from price_report ' + wheresql + ' ) group by cartype_name, crawl_date, source'
        cursor.execute(sql, values)
        result = cursor.fetchall()

        sql = 'select cartype_name, source, avg(msrp)-avg(price) differential ' \
              'from ( select cartype_name, source, price, msrp ' \
                        'from price_report ' + wheresql + ' ) group by cartype_name, source'

        # print type , sql.encode('gbk') , wheresql


        cursor.execute(sql, values)
        resultAll = cursor.fetchall()
        cursor.close()
        if len(result) <= 0:
            print '网站报价均值分析 未找到数据'.encode('gbk')
            return -1
        resultMap = {}
        for cell in result:
            if resultMap.has_key(cell[0]):
                (resultMap[cell[0]])[cell[1]+' '+cell[2]] = cell[3]
            else:
                resultMap[cell[0]] = {}
                (resultMap[cell[0]])[cell[1]+' '+cell[2]] = cell[3]
        resultAllMap = {}
        for cell in resultAll:
            if resultAllMap.has_key(cell[0]):
                (resultAllMap[cell[0]])[cell[1]] = cell[2]
            else:
                resultAllMap[cell[0]] = {}
                (resultAllMap[cell[0]])[cell[1]] = cell[2]

        rowNum += 1
        for cartype in self.sortCartype(conn, series):
            if resultMap.has_key(cartype):
                categoryMap = resultMap[cartype]
                length = 0
                for rowValue in settings.websiteDealer:
                    row = [rowValue]
                    hasData = False
                    for colValue in dateList:
                        categoryMapKey = rowValue + ' ' + colValue.strftime('%Y-%m-%d')
                        if categoryMap.has_key(categoryMapKey):
                            row.append(categoryMap[categoryMapKey])
                            hasData = True
                        else:
                            row.append('')
                    if hasData:
                        row.append(resultAllMap[cartype][rowValue])
                        self.setRow(colNum+1, rowNum, row, Style(), colHeight['content'])
                        length += 1
                        rowNum += 1
                self.mergeColCell(cartype, colNum, rowNum-length, length, Style())

        return rowNum

    def createAnalysisOfAreaPrice(self, conn, series, dateList, colNum, rowNum, type):
        colHeight = {'title': 21, 'thead': 13.5, 'content': 13.5}

        if type == 'all':
            title_end = ''
        elif type == 'all_autohome':
            title_end = '(汽车之家与易车)'
        self.mergeRowCell('销售大区报价分析'+title_end, colNum, rowNum, len(dateList)+3, colHeight['title'], Style('标题'))

        rowNum += 1
        self.mergeRowCell('报价日期 日期', colNum+2, rowNum, len(dateList)+1, colHeight['title'], Style('行标题2'))

        rowNum += 1
        thead = ['车型', '销售大区']
        for i in range(len(dateList)):
            thead.append(dateList[i].strftime('%m月%d日'))
        thead.append('总计')
        self.setRow(colNum, rowNum, thead, Style(), colHeight['thead'])

        cursor = conn.cursor()
        values = [series]
        wheresql = ' where carseries_name = :1  and IS_4S != \'#N/A\' '
        wheresql += ' and ( '
        for i in range(len(dateList)):
            if i != 0:
                wheresql += ' or '
            values.append(dateList[i].strftime('%Y-%m-%d'))
            wheresql += ' crawl_date = :1 '
        wheresql += ' ) '
        if type == 'all_autohome':
            wheresql += ' and ( SOURCE like \'汽车之家%\' or  SOURCE like \'易车%\' ) '
        sql = 'select cartype_name, sales_area, crawl_date, avg(msrp)-avg(price) differential ' \
              'from ( select cartype_name, sales_area, price, msrp, crawl_date ' \
                        'from price_report ' + wheresql + ' ) group by cartype_name, crawl_date, sales_area'
        cursor.execute(sql, values)
        result = cursor.fetchall()
        sql = 'select cartype_name, sales_area, avg(msrp)-avg(price) differential ' \
              'from ( select cartype_name, sales_area, price, msrp ' \
                        'from price_report ' + wheresql + ' ) group by cartype_name, sales_area'
        cursor.execute(sql, values)
        resultAll = cursor.fetchall()
        cursor.close()
        if len(result) <= 0:
            print "销售大区报价分析 未找到数据".encode('gbk')
            return -1
        resultMap = {}
        for cell in result:
            if resultMap.has_key(cell[0]):
                (resultMap[cell[0]])[cell[1]+' '+cell[2]] = cell[3]
            else:
                resultMap[cell[0]] = {}
                (resultMap[cell[0]])[cell[1]+' '+cell[2]] = cell[3]
        resultAllMap = {}
        for cell in resultAll:
            if resultAllMap.has_key(cell[0]):
                (resultAllMap[cell[0]])[cell[1]] = cell[2]
            else:
                resultAllMap[cell[0]] = {}
                (resultAllMap[cell[0]])[cell[1]] = cell[2]

        rowNum += 1
        for cartype in self.sortCartype(conn, series):
            if resultMap.has_key(cartype):
                categoryMap = resultMap[cartype]
                #一共有雪佛兰1~8区和无大区分类数据，这里需要且只需要雪佛兰1~8区的数据
                self.mergeColCell(cartype, colNum, rowNum, 8, Style())
                for rowValue in settings.area:
                    row = [rowValue]
                    for colValue in dateList:
                        categoryMapKey = rowValue + ' ' + colValue.strftime('%Y-%m-%d')
                        if categoryMap.has_key(categoryMapKey):
                            row.append(categoryMap[categoryMapKey])
                        else:
                            row.append('')
                    if len(row) > 1:
                        row.append(resultAllMap[cartype][rowValue])
                        self.setRow(colNum+1, rowNum, row, Style(), colHeight['content'])
                        rowNum += 1

        return rowNum

    def createTableAnalysisOfProvincesOffer(self, conn, series, dateList, colNum, rowNum, type):
        cursor = conn.cursor()
        values = [series]
        wheresql = ' where carseries_name = :1 and IS_4S != \'#N/A\' '
        wheresql += ' and ( '
        for i in range(len(dateList)):
            if i != 0:
                wheresql += ' or '
            values.append(dateList[i].strftime('%Y-%m-%d'))
            wheresql += ' crawl_date = :1 '
        wheresql += ' ) '
        if type == 'all_autohome':
            wheresql += ' and ( SOURCE like \'汽车之家%\' or  SOURCE like \'易车%\' ) '
        sql = 'select DISTINCT cartype_name from price_report ' + wheresql
        cursor.execute(sql, values)
        result = cursor.fetchall()
        if len(result) <= 0:
            return -1
        cartypeList = []
        for cartype in self.sortCartype(conn, series):
            for i in range(len(result)):
                if cartype == result[i][0]:
                    cartypeList.append(cartype)
                    result.remove(result[i])
                    break

        colHeight = {'title': 21, 'thead': 20.25, 'content': 13.5}

        if type == 'all':
            title_end = ''
        elif type == 'all_autohome':
            title_end = '(汽车之家与易车)'
        self.mergeRowCell('销售大区报价分析'+title_end, colNum, rowNum,
                          len(cartypeList)+2, colHeight['title'], Style('标题'))

        rowNum += 1
        self.mergeRowCell('车型', colNum+2, rowNum, len(cartypeList), colHeight['title'], Style('行标题2'))

        rowNum += 1
        thead = ['销售大区', '省份']
        for i in range(len(cartypeList)):
            thead.append(cartypeList[i])
        self.setRow(colNum, rowNum, thead, Style(), colHeight['thead'])

        sql = 'select sales_area, province, cartype_name, avg(msrp)-avg(price) differential ' \
              'from ( select cartype_name, sales_area, province, price, msrp, crawl_date ' \
                        'from price_report ' + wheresql + ' ) group by cartype_name, sales_area, province'
        cursor.execute(sql, values)
        result = cursor.fetchall()
        cursor.close()
        if len(result) <= 0:
            print "销售大区报价分析 未找到数据".encode('gbk')
            return -1
        resultMap = {}
        for cell in result:
            if resultMap.has_key(cell[0]):
                (resultMap[cell[0]])[cell[1]+' '+cell[2]] = cell[3]
            else:
                resultMap[cell[0]] = {}
                (resultMap[cell[0]])[cell[1]+' '+cell[2]] = cell[3]

        rowNum += 1
        for category in settings.area:
            categoryMap = resultMap[category]
            length = 0
            for province in settings.province[category]:
                row = [province]
                for cartype in cartypeList:
                    categoryMapKey = province + ' ' + cartype
                    if categoryMap.has_key(categoryMapKey):
                        row.append(categoryMap[categoryMapKey])
                    else:
                        row.append('')
                if len(row) > 1:
                    self.setRow(colNum+1, rowNum, row, Style(), colHeight['content'])
                    length += 1
                    rowNum += 1
            self.mergeColCell(category, colNum, rowNum-length, length, Style())

        return rowNum

    def createTableAnalysisOfDetailedQuotation(self, conn, series, dateList, colNum, rowNum):
        colHeight = {'title': 13.5, 'thead': 13.5, 'content': 13.5}

        fieldnameList = settings.fieldList
        style = Style()
        style.setFont('宋体', 11, True, 'FFFFFF')
        style.setPatternFill('5B9BD5')
        self.setRow(colNum, rowNum, fieldnameList, style, colHeight['title'])

        fieldnameStr = ''
        for i in range(len(fieldnameList)):
            if len(fieldnameStr) > 0:
                fieldnameStr += ' , '
            fieldnameStr += fieldnameList[i]
        cursor = conn.cursor()
        values = [series]
        wheresql = ' where carseries_name = :1 '
        values.append(dateList[len(dateList)-1].strftime('%Y-%m-%d'))
        wheresql += ' and crawl_date = :1 '
        sql = 'select ' + fieldnameStr + ' from price_report ' + wheresql
        cursor.execute(sql, values)
        result = cursor.fetchall()
        cursor.close()
        if len(result) <= 0:
            print ('    没有' + dateList[len(dateList)-1].strftime('%Y-%m-%d') + '的数据').encode('gbk')
            return -1
        print '    写入'.encode('gbk')+str(len(result))+'行数据'.encode('gbk')
        odd_style = Style()
        odd_style.setFont('宋体', 11, False, '000000')
        even_style = Style()
        even_style.setFont('宋体', 11, False, '000000')
        even_style.setPatternFill('DDEBF7')

        rowNum += 1
        for row in result:
            if rowNum % 2 == 1:
                self.setRow(colNum, rowNum, row, odd_style, colHeight['content'])
                if rowNum%100 == 0 :
                    print '      第'.encode('gbk')+str(rowNum-1)+'数据'.encode('gbk')
            else:
                self.setRow(colNum, rowNum, row, even_style, colHeight['content'])
                if rowNum%100 == 0:
                    print '      第'.encode('gbk')+str(rowNum-1)+'数据'.encode('gbk')
            rowNum += 1
        return rowNum

    def sortCartype(self, conn, series):
        cartypeList = []
        cursor = conn.cursor()
        sql = 'select cartype_name from price_report where carseries_name = :1 group by cartype_name order by avg(msrp)'
        cursor.execute(sql, [series])
        result = cursor.fetchall()
        cursor.close()
        if len(result) <= 0:
            print "该车系无车型".encode('gbk')
            return cartypeList
        for cell in result:
            if settings.markValue[series].has_key(cell[0]):
                cartypeList.append(cell[0])
        return cartypeList

    def singalAnalysisOfNetworkOffer(self, series, colNum, markValues, type):
        if type == 'all':
            title = '4S店,直营店,卫星店 - 网站报价均值分析'
        elif type == 'null':
            title = '未匹配经销商名称 - 网站报价均值分析'
        elif type == 'all_autohome':
            title = '4S店,直营店,卫星店 - 网站报价均值分析(汽车之家与易车)'
        else:
            title = '未匹配经销商名称 - 网站报价均值分析(汽车之家与易车)'
        rowNum = self.findCellRowNumByValue(colNum, title)

        if rowNum != -1:
            rowNum += 3
            while self.sheet[self.parseCellID(colNum, rowNum)].value is not None:
                cartype = self.sheet[self.parseCellID(colNum, rowNum)].value.encode('utf-8')
                for i in range(colNum+2, self.sheet.max_column):
                    cell = self.sheet[self.parseCellID(i, rowNum)]
                    if (cell.value is not None) and settings.carseries[series] == '<' \
                            and cell.value < markValues[cartype]:
                        self.setCellBgColor(cell, settings.markvalueColor)
                    elif (cell.value is not None) and settings.carseries[series] == '>' \
                            and cell.value > markValues[cartype]:
                        self.setCellBgColor(cell, settings.markvalueColor)
                rowNum += 1

    def singalAnalysisOfAreaPrice(self, series, colNum, markValues, type):
        if type == 'all':
            title = '销售大区报价分析'
        elif type == 'all_autohome':
            title = '销售大区报价分析(汽车之家与易车)'
        rowNum = self.findCellRowNumByValue(colNum, title)

        if rowNum != -1:
            rowNum += 3
            while self.sheet[self.parseCellID(colNum, rowNum)].value is not None:
                cartype = self.sheet[self.parseCellID(colNum, rowNum)].value.encode('utf-8')
                for i in range(colNum+2, self.sheet.max_column):
                    cell = self.sheet[self.parseCellID(i, rowNum)]
                    if (cell.value is not None) and settings.carseries[series] == '<' \
                            and cell.value < markValues[cartype]:
                        self.setCellBgColor(cell, settings.markvalueColor)
                    elif (cell.value is not None) and settings.carseries[series] == '>' \
                            and cell.value > markValues[cartype]:
                        self.setCellBgColor(cell, settings.markvalueColor)
                rowNum += 1

    def singalAnalysisOfProvincesOffer(self, series, colNum, markValues, type):
        if type == 'all':
            title = '销售大区报价分析'
        elif type == 'all_autohome':
            title = '销售大区报价分析(汽车之家与易车)'
        rowNum = self.findCellRowNumByValue(colNum, title)

        if rowNum != -1:
            rowNum += 2
            colNum += 2
            while self.sheet[self.parseCellID(colNum, rowNum)].value is not None:
                cartype = self.sheet[self.parseCellID(colNum, rowNum)].value.encode('utf-8')
                for i in range(34):
                    cell = self.sheet[self.parseCellID(colNum, rowNum+1+i)]
                    if (cell.value is not None) and settings.carseries[series] == '<' \
                            and cell.value < markValues[cartype]:
                        self.setCellBgColor(cell, settings.markvalueColor)
                    elif (cell.value is not None) and settings.carseries[series] == '>' \
                            and cell.value > markValues[cartype]:
                        self.setCellBgColor(cell, settings.markvalueColor)
                colNum += 1

    #寻找第colNum列中value出现的行数
    def findCellRowNumByValue(self, colNum, value):
        rowLength = self.sheet.max_row
        rowNum = -1
        for i in range(rowLength):
            if self.sheet[self.parseCellID(colNum, i+1)].value == value:
                rowNum = i+1
                break
        return rowNum

    #设置sheet页中每列的宽度，在xlsxFormatSetting的width中配置
    def setRowWidth(self):
        key = self.sheet.title.encode('utf-8')
        if settings.width.has_key(key):
            width = settings.width[key]
        else:
            width = [8.38]
        for i in range(len(width)):
            self.sheet.column_dimensions[self.parseCellRowID(i+1)].width = width[i]

    #合并一行的单元格
    #cellValue 单元格的值
    #colNum int 起始单元格行号
    #rowNum int 起始单元格列号
    #length int 合并单元格的长度
    #height int 单元格高度
    #style Style 单元格的格式
    def mergeRowCell(self, cellValue, colNum, rowNum, length, height, style):
        self.sheet.row_dimensions[rowNum].height = height
        self.sheet.merge_cells(self.parseCellID(colNum, rowNum)+':'+self.parseCellID(colNum+(length-1), rowNum))
        for i in range(length):
            cell = self.sheet[self.parseCellID(colNum+i, rowNum)]
            self.setCell(cell, cellValue, style)

    #合并一列的单元格
    #cellValue 单元格的值
    #colNum int 起始单元格行号
    #rowNum int 起始单元格列号
    #length int 合并单元格的长度
    #height int 单元格高度
    #style Style 单元格的格式
    def mergeColCell(self, cellValue, colNum, rowNum, length, style):
        self.sheet.merge_cells(self.parseCellID(colNum, rowNum)+':'+self.parseCellID(colNum, rowNum+(length-1)))
        for i in range(length):
            cell = self.sheet[self.parseCellID(colNum, rowNum+i)]
            self.setCell(cell, cellValue, style)

    #设置一行单元格(单元格使用统一格式)
    #colNum int 起始单元格行号
    #rowNum int 起始单元格列号
    #row 列表 单元格的值
    #style Style 单元格的格式
    #height int 高度
    def setRow(self, colNum, rowNum, row, style, height):
        self.sheet.row_dimensions[rowNum].height = height
        for i in range(len(row)):
            cell = self.sheet[self.parseCellID(colNum+i, rowNum)]
            self.setCell(cell, row[i], style)

    #修改一行单元格的背景色
    #colNum int 起始单元格行号
    #rowNum int 起始单元格列号
    #row 列表 单元格的值
    #color 六位字符串 表示十六进制下的颜色，如'FFFFFF'
    def setRowBgColor(self, colNum, rowNum, row, color):
        for i in range(len(row)):
            cell = self.sheet[self.parseCellID(colNum+i, rowNum)]
            self.setCellBgColor(cell, color)

    #设置单元格
    #cell Cell 单元格
    #cellValue int或str 单元格的值
    #style Style 单元格的风格
    def setCell(self, cell, cellValue, style):
        cell.value = cellValue
        cell.font = Font(name=style.font['family'],
                         size=style.font['size'],
                         bold=style.font['blod'],
                         color=style.font['color'])
        cell.fill = PatternFill(patternType=style.patternFill['patternType'],
                                start_color=style.patternFill['color'],
                                end_color=style.patternFill['color'])
        cell.alignment = Alignment(horizontal=style.alignment['horizontal'],
                                   vertical=style.alignment['vertical'],
                                   wrap_text=style.alignment['wrap_text'])
        cell.border = Border(left=Side(border_style=style.border['style'], color=style.border['color']),
                             right=Side(border_style=style.border['style'], color=style.border['color']),
                             top=Side(border_style=style.border['style'],  color=style.border['color']),
                             bottom=Side(border_style=style.border['style'], color=style.border['color']))
        cell.number_format = style.number_format['number']

    #单元格字体加粗
    def setCellBold(self, cell):
        font = cell.font
        cell.font = Font(name=font.name,
                         size=font.size,
                         bold=True,
                         color=font.color)

    #单元格字体改为宋体
    def setCellSongTi(self, cell):
        font = cell.font
        cell.font = Font(name='宋体',
                         size=font.size,
                         bold=font.bold,
                         color=font.color)

    #单元格背景颜色修改
    #cell Cell 单元格
    #color 6位字符串 十六进制下的颜色
    def setCellBgColor(self, cell, color):
        cell.fill = PatternFill(patternType='solid',
                                start_color='FF'+color,
                                end_color='FF'+color)

    #将列号、行号转变为excel单元格序列号，如 1,1 转变为 A1
    #colNum int 列号
    #rowNum int 行号
    def parseCellID(self, colNum, rowNum):
        if isinstance(colNum, int) and isinstance(rowNum, int):
            id = ''
            while colNum >= 1:
                id = chr((colNum-1) % 26 + 65) + id
                colNum = (colNum-1) / 26
            id += str(rowNum)
            return id
        else:
            return None

    #将列号转变为excel列，如 1 转变为 A
    #colNum int 列号
    def parseCellRowID(self, colNum):
        if isinstance(colNum, int):
            id = ''
            while colNum >= 1:
                id = chr((colNum-1) % 26 + 65) + id
                colNum = (colNum-1) / 26
            return id
        else:
            return None

if __name__ == '__main__':
    util = XlsxUtil(None)
    list = util.sortCartype(
        cx_Oracle.connect('admin/admin@10.0.0.233:15233/ORCL'), '全新科鲁兹'
    )
    for cell in list:
        print cell.encode('gbk')
