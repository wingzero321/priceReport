# -*- coding: utf-8 -*-

import sys
import os
import cx_Oracle
import time
import datetime
import xlsxFormatSetting as settings
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from xlsxUtil import *

os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'

reload(sys)
sys.setdefaultencoding('utf8')
reload(sys)


def createReportSpecification(sheet, conn, series, dateList):
    print "  正在生成 报表说明".encode('gbk')
    util = XlsxUtil(sheet)
    colNum = 2
    rowNum = 0
    util.setRowWidth()

    rowNum += 2
    rowNum = util.createReportInstruction(dateList, colNum, rowNum)

    rowNum += 4
    rowNum = util.createReportParameterSetting(conn, series, dateList, colNum, rowNum)

    print "  生成成功 报表说明".encode('gbk')


def createAnalysisOfWebsiteDealer(sheet, conn, series, dateList):
    print "  正在生成 网站经销商数量变化分析".encode('gbk')
    util = XlsxUtil(sheet)
    rowNum = 0
    colNum = 2
    util.setRowWidth()

    rowNum += 2
    rowNum = util.createReportsAnalysisOfWebsiteDealer(conn, series, dateList, colNum, rowNum, 'all')

    rowNum += 10
    rowNum = util.createReportsAnalysisOfWebsiteDealer(conn, series, dateList, colNum, rowNum, 'null')

    print "  生成成功 网站经销商数量变化分析".encode('gbk')


def createAnalysisOfNetworkOffer(sheet, conn, series, dateList):
    print "  正在生成 网站报价均值分析".encode('gbk')
    util = XlsxUtil(sheet)
    rowNum = 0
    colNum = 2
    util.setRowWidth()

    rowNum += 2
    rowNum = util.createAnalysisOfNetworkOffer(conn, series, dateList, colNum, rowNum, 'all')

    rowNum += 15
    rowNum = util.createAnalysisOfNetworkOffer(conn, series, dateList, colNum, rowNum, 'null')

    rowNum += 15
    rowNum = util.createAnalysisOfNetworkOffer(conn, series, dateList, colNum, rowNum, 'all_autohome')

    rowNum += 15
    rowNum = util.createAnalysisOfNetworkOffer(conn, series, dateList, colNum, rowNum, 'nul_autohome')

    print "  生成成功 网站报价均值分析".encode('gbk')


def createAnalysisOfAreaPrice(sheet, conn, series, dateList):
    print "  正在生成 大区报价分析".encode('gbk')
    util = XlsxUtil(sheet)
    rowNum = 0
    colNum = 2
    util.setRowWidth()

    rowNum += 2
    rowNum = util.createAnalysisOfAreaPrice(conn, series, dateList, colNum, rowNum, 'all')

    rowNum += 15
    rowNum = util.createAnalysisOfAreaPrice(conn, series, dateList, colNum, rowNum, 'all_autohome')

    print "  生成成功 大区报价分析".encode('gbk')


def createTableAnalysisOfProvincesOffer(sheet, conn, series, dateList):
    print "  正在生成 省份报价分析".encode('gbk')
    util = XlsxUtil(sheet)
    rowNum = 0
    colNum = 2
    util.setRowWidth()

    rowNum += 2
    rowNum = util.createTableAnalysisOfProvincesOffer(conn, series, dateList, colNum, rowNum, 'all')

    rowNum += 15
    rowNum = util.createTableAnalysisOfProvincesOffer(conn, series, dateList, colNum, rowNum, 'all_autohome')

    print "  生成成功 省份报价分析".encode('gbk')


def createTableAnalysisOfDetailedQuotation(sheet, conn, series, dateList):
    print "  正在生成 报价详细".encode('gbk')
    util = XlsxUtil(sheet)
    rowNum = 0
    colNum = 1
    util.setRowWidth()

    rowNum += 1
    rowNum = util.createTableAnalysisOfDetailedQuotation(conn, series, dateList, colNum, rowNum)

    print "  生成成功 报价详细".encode('gbk')


def createReport(conn, series, dateList):
    # 新建一个workbook
    wb = Workbook()
    # 默认sheet"报表说明"
    sheet = wb.worksheets[0]
    sheet.title = u'报表说明'
    createReportSpecification(sheet, conn, series, dateList)
    # 创建sheet"网站经销商数量变化分析"
    sheet = wb.create_sheet(u'网站经销商数量变化分析', 1)
    createAnalysisOfWebsiteDealer(sheet, conn, series, dateList)
    # 创建sheet"网站报价均值分析  "
    sheet = wb.create_sheet(u'网站报价均值分析', 2)
    createAnalysisOfNetworkOffer(sheet, conn, series, dateList)
    # 创建sheet"大区报价分析
    sheet = wb.create_sheet(u'大区报价分析', 3)
    createAnalysisOfAreaPrice(sheet, conn, series, dateList)
    # 创建sheet"省份报价分析"
    sheet = wb.create_sheet(u'省份报价分析', 4)
    createTableAnalysisOfProvincesOffer(sheet, conn, series, dateList)
    # 创建sheet"报价详细"
    sheet = wb.create_sheet(u'报价详细', 5)
    createTableAnalysisOfDetailedQuotation(sheet, conn, series, dateList)
    file_dir = settings.filePath + '\\report\\' + series.encode('gbk') + '-报价日报_'.encode('gbk') + \
               dateList[len(dateList) - 1].strftime('%Y-%m-%d') + '.xlsx'
    # 保存文件
    ew = ExcelWriter(workbook=wb)
    ew.save(filename=file_dir)


if __name__ == '__main__':
    conn = cx_Oracle.connect('admin/admin@10.0.0.233:15233/ORCL')
    print conn

    try:
        dateList = []
        if len(sys.argv) == 1:
            localtime = time.strftime('%Y-%m-%d', time.localtime(time.time()))
            startTime = datetime.datetime.strptime(localtime, '%Y-%m-%d')
            startTime += datetime.timedelta(days=-7)
            endTime = time.mktime(time.strptime(localtime, '%Y-%m-%d'))
            while time.mktime(startTime.timetuple()) <= endTime:
                dateList.append(startTime)
                startTime += datetime.timedelta(days=1)
        elif len(sys.argv) == 3:
            startTime = datetime.datetime.strptime(sys.argv[1], '%Y-%m-%d')
            endTime = time.mktime(time.strptime(sys.argv[2], '%Y-%m-%d'))
            while time.mktime(startTime.timetuple()) <= endTime:
                dateList.append(startTime)
                startTime += datetime.timedelta(days=1)
        elif len(sys.argv) == 2:
            dateList.append(datetime.datetime.strptime(sys.argv[1], '%Y-%m-%d'))

        print len(settings.carseries)

        for series in settings.carseries:
            print series.decode('UTF8').encode('gbk')
            createReport(conn, series, dateList)
            print '生成报表成功'.encode('gbk')
    finally:
        conn.close()
